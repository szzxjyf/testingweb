from openpyxl import load_workbook

file_path = r"C:\Users\...\your_file.xlsx"

wb = load_workbook(file_path, read_only=True, data_only=True)

# 获取第2个sheet名称
sheet_name = wb.sheetnames[1]
ws = wb[sheet_name]

print(f"Sheet name: {sheet_name}\n")

for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    print(row)
    if i >= 10:
        break

wb.close()
